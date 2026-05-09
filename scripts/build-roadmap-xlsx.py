#!/usr/bin/env python3
"""
Regenerate docs/roadmap.xlsx from docs/roadmap.csv.

The CSV is the source of truth (see docs/roadmap.md). The xlsx is a presentation
artifact for stakeholders who prefer Excel / Numbers / Sheets. This script applies a
fixed style:
  - bold + frozen header row, auto-filter across the whole table
  - light-green / yellow / orange backgrounds for horizons v1.0 / v1.1 / v2.0,
    gray for out-of-scope rows
  - status column tinted by value (planned/in-progress/done/deferred/dropped)
  - landscape print setup, header repeated on every printed page

Run from the repo root in a venv (PEP 668 blocks system-Python pip):

    python3 -m venv /tmp/xlsx-venv && /tmp/xlsx-venv/bin/pip install -q openpyxl
    /tmp/xlsx-venv/bin/python scripts/build-roadmap-xlsx.py
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = REPO_ROOT / "docs" / "roadmap.csv"
XLSX_PATH = REPO_ROOT / "docs" / "roadmap.xlsx"


# Color palette per horizon (the leftmost column of each row).
HORIZON_FILL = {
    "v1.0": PatternFill("solid", fgColor="FFE2EFDA"),  # light green — closest in time
    "v1.1": PatternFill("solid", fgColor="FFFFF2CC"),  # light yellow — mid horizon
    "v2.0": PatternFill("solid", fgColor="FFFCE4D6"),  # light orange — far horizon
    "oos":  PatternFill("solid", fgColor="FFEDEDED"),  # gray — out of scope
}

# Color palette per status.
STATUS_FILL = {
    "planned":     PatternFill("solid", fgColor="FFFFFFFF"),
    "in-progress": PatternFill("solid", fgColor="FFFFE699"),
    "done":        PatternFill("solid", fgColor="FFC6EFCE"),
    "deferred":    PatternFill("solid", fgColor="FFEDEDED"),
    "dropped":     PatternFill("solid", fgColor="FFD9D9D9"),
}

# Column widths tuned for the actual content. Indexed by Excel letter.
COLUMN_WIDTHS = {
    "A": 9,    # horizon
    "B": 12,   # target
    "C": 18,   # lane
    "D": 32,   # item
    "E": 70,   # deliverable (longest text)
    "F": 24,   # dependency
    "G": 7,    # effort
    "H": 26,   # owner
    "I": 12,   # status
    "J": 18,   # exit_criteria_link
}


def build() -> None:
    if not CSV_PATH.exists():
        sys.exit(f"error: {CSV_PATH} not found")

    with CSV_PATH.open(newline="") as f:
        rows = list(csv.reader(f))

    if not rows:
        sys.exit(f"error: {CSV_PATH} is empty")

    header = rows[0]
    expected_cols = len(header)
    bad = [i for i, r in enumerate(rows) if len(r) != expected_cols]
    if bad:
        sys.exit(f"error: rows with wrong column count: {bad}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Roadmap"
    ws.append(header)

    # Header style.
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx in range(1, expected_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Body — data rows with horizon + status tinting and a thin border around every cell.
    thin_border = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for r_idx, row in enumerate(rows[1:], start=2):
        horizon = row[0]
        status = row[8]
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            if c_idx == 1:
                cell.fill = HORIZON_FILL.get(horizon, PatternFill())
            elif c_idx == 9:
                cell.fill = STATUS_FILL.get(status, PatternFill())

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(expected_cols)}{len(rows)}"
    ws.freeze_panes = "A2"

    # Wrapped-text rows need extra height; header gets a tighter one.
    for r in range(2, len(rows) + 1):
        ws.row_dimensions[r].height = 38
    ws.row_dimensions[1].height = 28

    # Print setup — landscape, fit-to-width, repeat header on every printed page.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH.relative_to(REPO_ROOT)} ({len(rows) - 1} rows)")


if __name__ == "__main__":
    build()
